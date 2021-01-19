"""DGT specific utils"""

import time
from collections import defaultdict
from datetime import date

import arrow
from naf import DB as NAF
from openpyxl import Workbook, load_workbook
from progressist import ProgressBar

from egapro import constants, db, models
from egapro.solen import ExcelData, RowProcessor
from egapro.utils import flatten, remove_one_year
from egapro.schema.legacy import from_legacy


AGES = {
    ":29": "30",
    "30:39": "30-39",
    "40:49": "40-49",
    "50:": "50",
}
EFFECTIF = {"50:250": "50 à 250", "251:999": "251 à 999", "1000:": "1000 et plus"}


def truthy(val):
    return bool(val)


def falsy(val):
    return not truthy(val)


def isodatetime(val):
    if not val:
        return None
    # Excel doesn't know nothing about timezone, so let's transpose.
    when = arrow.get(val).to("Europe/Paris").naive
    return when


def isodate(val):
    if not val:
        return None
    return date.fromisoformat(val)


def code_naf(code):
    if not code:
        return None
    return f"{code} - {NAF[code]}"


async def get_headers_columns():
    """Return a tuple of lists of (header_names, column_names) that we want in the export."""
    try:
        num_coefficient = await db.declaration.fetchval(
            "SELECT "
            "jsonb_array_length(data->'indicateurs'->'rémunérations'->'catégories') AS length "
            "FROM declaration WHERE data->'indicateurs'->'rémunérations' ? 'catégories' "
            "ORDER BY length DESC LIMIT 1;"
        )
    except db.NoData:
        num_coefficient = 0
    interesting_cols = (
        [
            ("source", "source"),
            ("URL_declaration", "URL_declaration"),  # Built from /data/id, see below
            ("Date_declaration", "déclaration.date", isodatetime),
            ("Date_modification", "modified_at", isodatetime),
            ("Email_declarant", "déclarant.email"),
            ("Nom", "déclarant.nom"),
            ("Prenom", "déclarant.prénom"),
            ("Telephone", "déclarant.téléphone"),
            ("Region", "entreprise.région", constants.REGIONS.get),
            ("Departement", "entreprise.département", constants.DEPARTEMENTS.get),
            ("Adresse", "entreprise.adresse"),
            ("CP", "entreprise.code_postal"),
            ("Commune", "entreprise.commune"),
            (
                "Annee_indicateurs",
                "déclaration.année_indicateurs",
            ),
            ("Date_debut_periode", "déclaration.début_période_référence"),
            ("Date_fin_periode", "déclaration.fin_période_référence"),
            ("Structure", "entreprise.structure"),
            ("Tranche_effectif", "entreprise.effectif.tranche", EFFECTIF.get),
            ("Nb_salaries", "entreprise.effectif.total"),
            ("Nom_Entreprise", "entreprise.raison_sociale"),
            ("SIREN", "entreprise.siren"),
            ("Code_NAF", "entreprise.code_naf", code_naf),
            ("Nom_UES", "entreprise.ues.nom"),
            # Inclure entreprise déclarante
            ("Nb_ets_UES", "entreprise.nombre_ues"),
            (
                "Date_publication",
                "déclaration.publication.date",
                isodate,
            ),
            ("Site_internet_publication", "déclaration.publication.url"),
            ("Modalites_publication", "déclaration.publication.modalités"),
            ("Indic1_calculable", "indicateurs.rémunérations.non_calculable_bool"),
            (
                "Indic1_motif_non_calculable",
                "indicateurs.rémunérations.non_calculable",
            ),
            ("Indic1_modalite_calcul", "indicateurs.rémunérations.mode"),
            (
                "Indic1_date_consult_CSE",
                "indicateurs.rémunérations.date_consultation_cse",
                isodate,
            ),
            ("Indic1_nb_coef_niv", "indicateurs.rémunérations.len_categories"),
            ("Indic1_Ouv", "indicateurs.rémunérations.Indic1_Ouv"),
            ("Indic1_Emp", "indicateurs.rémunérations.Indic1_Emp"),
            ("Indic1_TAM", "indicateurs.rémunérations.Indic1_TAM"),
            ("Indic1_IC", "indicateurs.rémunérations.Indic1_IC"),
        ]
        + [
            (
                f"Indic1_Niv{index_coef}",
                f"indicateurs.rémunérations.catégories.{index_coef}",
            )
            for index_coef in range(num_coefficient)
        ]
        + [
            ("Indic1_resultat", "indicateurs.rémunérations.résultat"),
            (
                "Indic1_population_favorable",
                "indicateurs.rémunérations.population_favorable",
            ),
            ("Indic2_calculable", "indicateurs.augmentations.non_calculable_bool"),
            (
                "Indic2_motif_non_calculable",
                "indicateurs.augmentations.non_calculable",
            ),
        ]
        + [
            (
                f"Indic2_{CSP}",
                f"indicateurs.augmentations.catégories.{index_csp}",
            )
            for (index_csp, CSP) in enumerate(["Ouv", "Emp", "TAM", "IC"])
        ]
        + [
            ("Indic2_resultat", "indicateurs.augmentations.résultat"),
            (
                "Indic2_population_favorable",
                "indicateurs.augmentations.population_favorable",
            ),
            ("Indic3_calculable", "indicateurs.promotions.non_calculable_bool"),
            ("Indic3_motif_non_calculable", "indicateurs.promotions.non_calculable"),
        ]
        + [
            (
                f"Indic3_{CSP}",
                f"indicateurs.promotions.catégories.{index_csp}",
            )
            for (index_csp, CSP) in enumerate(["Ouv", "Emp", "TAM", "IC"])
        ]
        + [
            ("Indic3_resultat", "indicateurs.promotions.résultat"),
            (
                "Indic3_population_favorable",
                "indicateurs.promotions.population_favorable",
            ),
            (
                "Indic2et3_calculable",
                "indicateurs.augmentations_et_promotions.non_calculable_bool",
            ),
            (
                "Indic2et3_motif_non_calculable",
                "indicateurs.augmentations_et_promotions.non_calculable",
            ),
            (
                "Indic2et3_resultat_pourcent",
                "indicateurs.augmentations_et_promotions.résultat",
            ),
            (
                "Indic2et3_resultat_nb_sal",
                "indicateurs.augmentations_et_promotions.résultat_nombre_salariés",
            ),
            (
                "Indic2et3_population_favorable",
                "indicateurs.augmentations_et_promotions.population_favorable",
            ),
            ("Indic4_calculable", "indicateurs.congés_maternité.non_calculable_bool"),
            (
                "Indic4_motif_non_calculable",
                "indicateurs.congés_maternité.non_calculable",
            ),
            ("Indic4_resultat", "indicateurs.congés_maternité.résultat"),
            ("Indic5_resultat", "indicateurs.hautes_rémunérations.résultat"),
            (
                "Indic5_sexe_sur_represente",
                "indicateurs.hautes_rémunérations.population_favorable",
            ),
            ("Indicateur_1", "indicateurs.rémunérations.note"),
            ("Indicateur_2", "indicateurs.augmentations.note"),
            ("Indicateur_3", "indicateurs.promotions.note"),
            ("Indicateur_2et3", "indicateurs.augmentations_et_promotions.note"),
            (
                "Indicateur_2et3_PourCent",
                "indicateurs.augmentations_et_promotions.note_en_pourcentage",
            ),
            (
                "Indicateur_2et3_ParSal",
                "indicateurs.augmentations_et_promotions.note_nombre_salariés",
            ),
            ("Indicateur_4", "indicateurs.congés_maternité.note"),
            ("Indicateur_5", "indicateurs.hautes_rémunérations.note"),
            ("Nombre_total_points obtenus", "déclaration.points"),
            (
                "Nombre_total_points_pouvant_etre_obtenus",
                "déclaration.points_calculables",
            ),
            ("Resultat_final_sur_100_points", "déclaration.index"),
            ("Mesures_correction", "déclaration.mesures_correctives"),
        ]
    )
    headers = []
    columns = []
    for header, column, *fmt in interesting_cols:
        headers.append(header)
        columns.append((column, fmt[0] if fmt else lambda x: x))
    return (headers, columns)


async def as_xlsx(max_rows=None, debug=False):
    """Export des données au format souhaité par la DGT.

    :max_rows:          Max number of rows to process.
    :debug:             Turn on debug to be able to read the generated Workbook
    """
    print("Reading from DB")
    records = await db.declaration.completed()
    print("Flattening JSON")
    if max_rows:
        records = records[:max_rows]
    wb = Workbook(write_only=not debug)
    ws = wb.create_sheet()
    ws.title = "BDD REPONDANTS"
    wb.active = ws
    ws_ues = wb.create_sheet()
    ws_ues.title = "BDD UES détail entreprises"
    ws_ues.append(
        [
            "Annee_indicateurs",
            "Region",
            "Departement",
            "Adresse",
            "CP",
            "Commune",
            "Tranche_effectif",
            "Nom_UES",
            "Siren_entreprise_declarante",
            "Nom_entreprise_declarante",
            "Nom_entreprise",
            "Siren",
        ]
    )
    headers, columns = await get_headers_columns()
    ws.append(headers)
    bar = ProgressBar(prefix="Computing", total=len(records))
    for record in bar.iter(records):
        data = record.data
        if not data:
            continue
        if "déclaration" not in data:  # Legacy schema
            from_legacy(data)
        ues_data(ws_ues, data)
        data = prepare_record(data)
        data["modified_at"] = record["modified_at"]
        ws.append([fmt(data.get(c)) for c, fmt in columns])
    return wb


def ues_data(sheet, data):
    data = models.Data(data)
    region = constants.REGIONS.get(data.path("entreprise.région"))
    departement = constants.DEPARTEMENTS.get(data.path("entreprise.département"))
    adresse = data.path("entreprise.adresse")
    cp = data.path("entreprise.code_postal")
    commune = data.path("entreprise.commune")
    tranche = EFFECTIF.get(data.path("entreprise.effectif.tranche"))
    nom = data.path("entreprise.ues.nom")
    for ues in data.path("entreprise.ues.entreprises") or []:
        sheet.append(
            [
                data.year,
                region,
                departement,
                adresse,
                cp,
                commune,
                tranche,
                nom,
                data.siren,
                data.company,
                ues["raison_sociale"],
                ues["siren"],
            ]
        )


def prepare_record(data):

    # Before flattening.
    data["URL_declaration"] = f"'https://index-egapro.travail.gouv.fr/{data.uri}"
    effectif = data["entreprise"]["effectif"]["tranche"]
    prepare_entreprise(data["entreprise"])
    prepare_declaration(data["déclaration"])
    prepare_remunerations(data["indicateurs"]["rémunérations"])
    prepare_conges_maternite(data["indicateurs"]["congés_maternité"])
    if effectif == "50:250":
        prepare_augmentations_et_promotions(
            data["indicateurs"]["augmentations_et_promotions"]
        )
    else:
        prepare_augmentations(data["indicateurs"]["augmentations"])
        prepare_promotions(data["indicateurs"]["promotions"])

    return flatten(data, flatten_lists=True)


def prepare_entreprise(data):
    nombre_ues = len(data.get("ues", {}).get("entreprises", []))
    data["structure"] = (
        "Unité Economique et Sociale (UES)" if nombre_ues else "Entreprise"
    )
    data["nombre_ues"] = nombre_ues or None


def prepare_declaration(data):
    data["début_période_référence"] = remove_one_year(
        date.fromisoformat(data["fin_période_référence"])
    )
    data["fin_période_référence"] = date.fromisoformat(data["fin_période_référence"])


def prepare_remunerations(data):
    try:
        indic1_categories = data["catégories"]
    except KeyError:
        indic1_categories = []
    indic1_nv_niveaux = len(indic1_categories) or None
    indic1_mode = data.get("mode")
    data["len_categories"] = indic1_nv_niveaux if indic1_mode != "csp" else None
    calculable = not data.get("non_calculable")
    if calculable:
        # DGT want data to be in different columns whether its csp or any coef.
        csp_names = ["Ouv", "Emp", "TAM", "IC"]
        for idx, category in enumerate(indic1_categories):
            tranches = category.get("tranches", {})
            key = f"catégories.{idx}"
            if indic1_mode == "csp":
                key = f"Indic1_{csp_names[idx]}"
            values = [
                tranches.get(":29"),
                tranches.get("30:39"),
                tranches.get("40:49"),
                tranches.get("50:"),
            ]
            # Prevent "-0.0" or "0.0" or "12.0" as str representation
            values = [int(v) if v is not None and v % 1 == 0 else v for v in values]
            values = [str(round(v, 2) + 0) if v is not None else "nc" for v in values]
            data[key] = ";".join(values)
    else:
        data["note"] = "nc"
    data["non_calculable_bool"] = calculable


def prepare_promotions(data):
    calculable = not data.get("non_calculable")
    data["non_calculable_bool"] = calculable
    if not calculable:
        data["note"] = "nc"


def prepare_augmentations(data):
    calculable = not data.get("non_calculable")
    data["non_calculable_bool"] = calculable
    if not calculable:
        data["note"] = "nc"


def prepare_augmentations_et_promotions(data):
    calculable = not data.get("non_calculable")
    data["non_calculable_bool"] = calculable
    if not calculable:
        data["note"] = "nc"


def prepare_conges_maternite(data):
    calculable = not data.get("non_calculable")
    data["non_calculable_bool"] = calculable
    if not calculable:
        data["note"] = "nc"


async def duplicates(current_export, *solen_data):  # pragma: no cover
    before = time.perf_counter()
    headers, columns = await get_headers_columns()
    columns = [c for c, _ in columns]
    reversed_headers = dict(zip(headers, columns))
    data = defaultdict(list)
    raw = list(
        load_workbook(current_export, read_only=True, data_only=True).active.values
    )
    timer, before = time.perf_counter() - before, time.perf_counter()
    print(f"Done reading current data ({timer})")
    own_headers = raw[0]
    year_idx = own_headers.index("Annee_indicateurs")
    siren_idx = own_headers.index("SIREN")
    for row in raw[1:]:
        if row[0].startswith("solen"):
            continue
        year = row[year_idx]
        siren = row[siren_idx]
        key = f"{year}.{siren}"
        # Align to current headers (which change according to data in DB)
        record = {
            reversed_headers[own_headers[i]]: row[i]
            for i in range(len(row))
            if own_headers[i] in reversed_headers
        }
        data[key].append(record)
    timer, before = time.perf_counter() - before, time.perf_counter()
    print(f"Done filtering current data ({timer})")
    for path in solen_data:
        _, year = path.stem.split("-")
        raw = ExcelData(path)
        for row in raw.repondants.values():
            record = flatten(
                RowProcessor(
                    year,
                    None,
                    row,
                ).run()["data"]
            )
            url = (
                f"'https://solen1.enquetes.social.gouv.fr/cgi-bin/HE/P?P={record['id']}"
            )
            record["URL_declaration"] = url
            siren = record["entreprise.siren"]
            year = record["déclaration.année_indicateurs"]
            key = f"{year}.{siren}"
            data[key].append(record)
    timer, before = time.perf_counter() - before, time.perf_counter()
    print(f"Done reading solen data: ({timer})")
    print(f"Unique entries: {len(data)}")
    duplicates = {k: v for k, v in data.items() if len(v) > 1}
    timer, before = time.perf_counter() - before, time.perf_counter()
    print(f"Done computing duplicates ({timer})")
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    wb.active = ws
    ws.append(headers)
    for entry in duplicates.values():
        for record in entry:
            ws.append([record.get(c) for c in columns])
    return wb
